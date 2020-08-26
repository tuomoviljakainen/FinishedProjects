# -*- coding: UTF-8 -*-
#!/usr/bin/python3
import random
import time
import openpyxl
import string
import sys
import timeit

# Stackkiin saadaan suurempi määrä dataa; scripti ei kaadu niin helposti
sys.setrecursionlimit(7500)

operaatiot = 0
print("Alkaa")

# Ensimmäisen alkion toteutustapa
def EnsimmainenAlkio(taulukko):
    global operaatiot
    # Taulukot, joihin alkioita tallennetaan kun niitä verrataan sarana-alkioon.
    # pienemmat-taulukkoon tallennetaan sarana-alkiota pienemmät alkiot.
    pienemmat = []
    # yhtasuuret-taulukkoon sarana-alkion kanssa yhtäsuuret alkiot.
    yhtasuuret = []
    # suuremmat-taulukkoon tallennetaan sarana-alkiota suuremmat alkiot.
    suuremmat = []

    # Jos taulukon koko on suurempi kuin yksi.
    if len(taulukko) > 1:
        # Valitaan sarana-alkioksi taulukon ensimmäinen solu 
        sarana_alkio = taulukko[0]
        # Käydään kaikki taulukon solut yksi kerrallaan läpi
        for x in taulukko:
            # Jos sarana-alkio on pienempi kuin solu, 
            # niin solu lisätään pienemmat-taulukkoon.
            if x < sarana_alkio:
                pienemmat.append(x)
                operaatiot += 1
            # Jos sarana-alkio on yhtä suuri kuin solu, 
            # niin solu lisätään yhtasuuret-taulukkoon.
            elif x == sarana_alkio:
                yhtasuuret.append(x)
                operaatiot += 1
            # Jos sarana-alkio on suurempi kuin solu, 
            # niin solu lisätään suuremmat-taulukkoon.
            else:
                suuremmat.append(x)
                operaatiot += 1
        # Kutsutaan pienemmat ja suuremmat taulukkoa rekursiivisesti. yhtasuuret-
        # taulukon solut ovat löytäneet paikkansa lopullisessa järjestyksessä.
        return EnsimmainenAlkio(pienemmat)+yhtasuuret+EnsimmainenAlkio(suuremmat)
    # Jos taulukon koko on pienempi kuin kaksi, se palautetaan.
    else:
        return taulukko

# Satunnaisen alkion toteutustapa
def SatunnainenAlkio(taulukko):
    global operaatiot
    # Taulukot, joihin alkioita tallennetaan kun niitä verrataan sarana-alkioon.
    # pienemmat-taulukkoon tallennetaan sarana-alkiota pienemmät alkiot.
    pienemmat = []
    # yhtasuuret-taulukkoon sarana-alkion kanssa yhtäsuuret alkiot.
    yhtasuuret = []
    # suuremmat-taulukkoon tallennetaan sarana-alkiota suuremmat alkiot.
    suuremmat = []

    # Jos taulukon koko on suurempi kuin yksi.
    if len(taulukko) > 1:
        # Valitaan sarana-alkio valitsemalla satunnainen luku taulukosta
        sarana_alkio = random.choice(taulukko)
        # Käydään kaikki taulukon solut yksi kerrallaan läpi
        for x in taulukko:
            # Jos sarana-alkio on pienempi kuin solu, 
            # niin solu lisätään pienemmat-taulukkoon.
            if x < sarana_alkio:
                pienemmat.append(x)
                operaatiot += 1
            # Jos sarana-alkio on yhtä suuri kuin solu, 
            # niin solu lisätään yhtasuuret-taulukkoon.
            elif x == sarana_alkio:
                yhtasuuret.append(x)
                operaatiot += 1
            # Jos sarana-alkio on suurempi kuin solu, 
            # niin solu lisätään suuremmat-taulukkoon.
            else:
                suuremmat.append(x)
                operaatiot += 1
        # Kutsutaan pienemmat ja suuremmat taulukkoa rekursiivisesti. yhtasuuret-
        # taulukon solut ovat löytäneet paikkansa lopullisessa järjestyksessä.
        return SatunnainenAlkio(pienemmat)+yhtasuuret+SatunnainenAlkio(suuremmat)
    # Jos taulukon koko on pienempi kuin kaksi, se palautetaan.
    else:
        return taulukko

# Keskimmäinen alkio kolmesta -toteutustapa
def KeskimmainenKolmestaAlkio(taulukko):
    global operaatiot
    # Taulukot, joihin alkioita tallennetaan kun niitä verrataan sarana-alkioon.
    # pienemmat-taulukkoon tallennetaan sarana-alkiota pienemmät alkiot.
    pienemmat = []
    # yhtasuuret-taulukkoon sarana-alkion kanssa yhtäsuuret alkiot.
    yhtasuuret = []
    # suuremmat-taulukkoon tallennetaan sarana-alkiota suuremmat alkiot.
    suuremmat = []

    # Jos taulukon koko on suurempi kuin yksi.
    if len(taulukko) > 1:
        # Valitaan 3 satunnaista lukua taulukosta ja niistä arvontaan keskimmäistä käytetään sarana-alkiona
        sarana_alkio = sorted(random.choices(taulukko, k=3))[1]
        # Käydään kaikki taulukon solut yksi kerrallaan läpi
        for x in taulukko:
            # Jos sarana-alkio on pienempi kuin solu, 
            # niin solu lisätään pienemmat-taulukkoon.
            if x < sarana_alkio:
                pienemmat.append(x)
                operaatiot += 1
            # Jos sarana-alkio on yhtä suuri kuin solu, 
            # niin solu lisätään yhtasuuret-taulukkoon.
            elif x == sarana_alkio:
                yhtasuuret.append(x)
                operaatiot += 1
            # Jos sarana-alkio on suurempi kuin solu, 
            # niin solu lisätään suuremmat-taulukkoon.
            else:
                suuremmat.append(x)
                operaatiot += 1
        # Kutsutaan pienemmat ja suuremmat taulukkoa rekursiivisesti. yhtasuuret-
        # taulukon solut ovat löytäneet paikkansa lopullisessa järjestyksessä.
        return SatunnainenAlkio(pienemmat)+yhtasuuret+SatunnainenAlkio(suuremmat)
    # Jos taulukon koko on pienempi kuin kaksi, se palautetaan.
    else:
        return taulukko

def laske(taulukko, otot, kolumni, solujenLkm):
    # Otetaan lajitteluita niin paljon kuin ottoja on määritelty
    for i in range(otot):
        global operaatiot
        operaatiot = 0
        numerot = []
        # Lisätään taulukkoon solu satunnaisesti 0 - 100 000 000 väliltä
        for x in range(solujenLkm):
            numero = random.randint(0,100000000)
            numerot.append(numero)

        # Syötetään taulukko laskettavaksi ensimmäisen alkion toteutustavalle ja lasketaan kauan aikaa lajitteluun menee
        aloitusAika = timeit.default_timer()
        EnsimmainenAlkio(numerot)
        kaytettyAika = timeit.default_timer() - aloitusAika
        # Tulostetaan tulokset
        print("Alkio määrä: " + str(otot))
        print("Taulukon koko: " + str(len(numerot)))
        print("\nEnsimmäinen alkio:")
        print("käytetty aika: ", kaytettyAika)
        print("Operaatioiden määrä: ", operaatiot)
        # Lisätään tulokset taulukkoon
        taulukko.cell(row=i+3, column=kolumni).value = operaatiot
        taulukko.cell(row=i+3, column=kolumni+1).value = kaytettyAika

        # Syötetään taulukko laskettavaksi satunnaisen alkion toteutustavalle ja lasketaan kauan aikaa lajitteluun menee
        operaatiot = 0
        aloitusAika = timeit.default_timer()
        SatunnainenAlkio(numerot)
        kaytettyAika = timeit.default_timer() - aloitusAika
        # Tulostetaan tulokset
        print("\nSatunnainen alkio:")
        print("käytetty aika: ", kaytettyAika)
        print("Operaatioiden määrä: ", operaatiot)
        # Lisätään tulokset taulukkoon
        taulukko.cell(row=i+3, column=kolumni+2).value = operaatiot
        taulukko.cell(row=i+3, column=kolumni+3).value = kaytettyAika

        # Syötetään taulukko laskettavaksi keskimmäinen alkio kolmesta -toteutustavalle ja lasketaan kauan aikaa lajitteluun menee
        operaatiot = 0
        aloitusAika = timeit.default_timer()
        KeskimmainenKolmestaAlkio(numerot)
        kaytettyAika = timeit.default_timer() - aloitusAika
        # Tulostetaan tulokset
        print("\nParas kolmesta alkio:")
        print("käytetty aika: ", kaytettyAika)
        print("Operaatioiden määrä: ", operaatiot)
        # Lisätään tulokset taulukkoon
        taulukko.cell(row=i+3, column=kolumni+4).value = operaatiot
        taulukko.cell(row=i+3, column=kolumni+5).value = kaytettyAika
        print("--------------------------------")


if __name__ == '__main__':
    
    # Määrittää kuinka monta otosta otetaan per alkiomäärä
    otot = 2

    # Lisätään dataa excel-taulukkoon
    TIEDOSTO = openpyxl.Workbook()
    taulukko = TIEDOSTO.active
    taulukko["A1"] = "Solujen määrä"
    taulukko["A2"] = 5
    taulukko["B1"] = "Ensimmäinen alkio sarana-alkiona"
    taulukko["B2"] = "Operaatiot"
    taulukko["C2"] = "Käytetty aika"
    taulukko["D1"] = "Satunnainen alkio sarana-alkiona"
    taulukko["D2"] = "Operaatiot"
    taulukko["E2"] = "Käytetty aika"
    taulukko["F1"] = "Sarana-alkioksi valittu keskimmäinen kolmesta"
    taulukko["F2"] = "Operaatiot"
    taulukko["G2"] = "Käytetty aika"

    # lasketaan 10 000 alkiolle keskiarvo
    laske(taulukko, otot, 2, 10000)

    taulukko["I1"] = "Solujen määrä"
    taulukko["I2"] = 10
    taulukko["J1"] = "Ensimmäinen alkio sarana-alkiona"
    taulukko["J2"] = "Operaatiot"
    taulukko["K2"] = "Käytetty aika"
    taulukko["L1"] = "Satunnainen alkio sarana-alkiona"
    taulukko["L2"] = "Operaatiot"
    taulukko["M2"] = "Käytetty aika"
    taulukko["N1"] = "Sarana-alkioksi valittu keskimmäinen kolmesta"
    taulukko["N2"] = "Operaatiot"
    taulukko["O2"] = "Käytetty aika"

    # lasketaan 100 000 alkiolle keskiarvo
    laske(taulukko, otot, 10, 100000)

    taulukko["Q1"] = "Solujen määrä"
    taulukko["Q2"] = 100
    taulukko["R1"] = "Ensimmäinen alkio sarana-alkiona"
    taulukko["R2"] = "Operaatiot"
    taulukko["S2"] = "Käytetty aika"
    taulukko["T1"] = "Satunnainen alkio sarana-alkiona"
    taulukko["T2"] = "Operaatiot"
    taulukko["U2"] = "Käytetty aika"
    taulukko["V1"] = "Sarana-alkioksi valittu keskimmäinen kolmesta"
    taulukko["V2"] = "Operaatiot"
    taulukko["W2"] = "Käytetty aika"
    
    # lasketaan 1 000 000 alkiolle keskiarvo
    laske(taulukko, otot, 18, 1000000)

    # Muutetaan sarakkeiden koko 30 excelissä
    for sarake in string.ascii_lowercase:
        taulukko.column_dimensions[sarake].width = 30
    # Keskitetään alkiot excelissä
    for rivinumero in range(1,otot+3):
        for sarake in taulukko[str(rivinumero)+":"+str(rivinumero)]:
            sarake.alignment = openpyxl.styles.Alignment(horizontal="center")
    # Excel-tiedoston nimi johon tulokset tallennetaan
    TIEDOSTO.save('Satunnaiset_numero.xlsx')
    print("Valmis!")